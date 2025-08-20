import random,sys,os,shutil,re,itertools
from datetime import datetime
from tabulate import tabulate
from termcolor import colored
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Constants
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__),'..'))
GENERATOR_PARAMETERS_FILE = os.path.join(BASE_DIR, "io","generator","input","parameters.txt")
GENERATOR_SPECIES_FILE = os.path.join(BASE_DIR, "io","generator","input","species.txt")
GENERATOR_FOOD_FILE = os.path.join(BASE_DIR,"io","generator","input","food.txt")
LAUNCHER_PARAMETERS_FILE = os.path.join(BASE_DIR, "io","launcher","input","parameters.txt")
KAUFFMAN_GENERATOR_PARAMETERS_FILE = os.path.join(BASE_DIR, "io","kauffmanGenerator","input","parameters.txt")
KAUFFMAN_GENERATOR_SPECIES_FILE = os.path.join(BASE_DIR, "io","kauffmanGenerator","input","species.txt")
MUTATOR_INPUT = os.path.join(BASE_DIR, "io","mutator","input")
MUTATOR_OUTPUT = os.path.join(BASE_DIR, "io","mutator","output")
MUTATOR_CHEMISTRY_FILE = os.path.join(MUTATOR_INPUT, "chemistry.txt")
MUTATOR_PARAMETERS_FILE = os.path.join(MUTATOR_INPUT, "parameters.txt")
MUTATOR_RULES_FILE = os.path.join(MUTATOR_INPUT, "rules.txt")
MUTATOR_SPECIES_FILE = os.path.join(MUTATOR_INPUT, "species.txt")
MUTATOR_OUTPUT_FILE = os.path.join(MUTATOR_OUTPUT, "uniqueReactions.txt")
MUTATOR_OUTPUT_RULES_FILE = os.path.join(MUTATOR_OUTPUT, "rules.txt")
MUTATOR_DIFFERENCE_FILE = os.path.join(MUTATOR_OUTPUT, "difference.txt")
EVOLVER_INPUT = os.path.join(BASE_DIR, "io","evolver","input")
EVOLVER_OUTPUT = os.path.join(BASE_DIR, "io","evolver","output")
EVOLVER_PARAMETERS_FILE = os.path.join(EVOLVER_INPUT, "parameters.txt")
EVOLVER_CHEMISTRY_PARAMETERS_FILE = os.path.join(EVOLVER_INPUT,"chemistry_info","parameters.txt")
EVOLVER_CHEMISTRY_RULES_FILE = os.path.join(EVOLVER_INPUT,"chemistry_info","rules.txt")
EVOLVER_CHEMISTRY_WITHOUT_CONTAINER_FILE = os.path.join(EVOLVER_INPUT,"chemistry_info","chemistryNoContainer.txt")
EVOLVER_CHEMISTRY_WITH_CONTAINER_FILE = os.path.join(EVOLVER_INPUT,"chemistry_info","chemistryContainer.txt")
VENTURI = os.path.join(BASE_DIR,"utils","venturi")
PITZALIS = os.path.join(BASE_DIR,"utils","pitzalis")
PITZALIS_CONTAINERS_RULES = os.path.join(PITZALIS,"src","Utils","sp_cat_membrana.txt")
PITZALIS_CONTAINERS_SPECIES_INPUT = os.path.join(PITZALIS,"src","Utils","chimica_in.txt")
PITZALIS_CONTAINERS_SPECIES_OUTPUT = os.path.join(PITZALIS,"src","Utils","chimica.txt")
PITZALIS_INPUT = os.path.join(PITZALIS,"input","chimica.txt")
PITZALIS_OUTPUT = os.path.join(PITZALIS,"out")
PITZALIS_TIME = os.path.join(PITZALIS, "src", "tempo.txt")
PITZALIS_QUANTITIES = os.path.join(PITZALIS, "src", "quantita.txt")
NUMERIC_START = re.compile(r'^\s*(\d+(\.\d*)?|\.\d+)')

# Utils functions
def readFile(path):
    f=open(path)
    file=f.readlines()
    f.close()
    return file

def getRandomValue():
    return random.randrange(sys.maxsize)

def copyFile(source,destination):
    shutil.copy(source,destination)

def formatFileForVenturi (path, chemistry=False):
    with open(path, "r") as f:
        lines = f.readlines()
    formattedFile = [l for l in lines if l.strip() != "NESSUNA REAZIONE GENERATA, SI CONSIGLIA DI MODIFICARE I PARAMETRI"]
    formattedFile = formattedFile[2:]
    with open(path, "w") as f:
        f.writelines(formattedFile)

def createDirectory(path):
    if not os.path.exists(path):
        os.makedirs(path)

def runVenturi(venturiPath):
    os.system(f'cd "{venturiPath}/CRST_src" && python ./Driver.py cta > nul 2>&1')

def monomerCombinations( monomers, outerRadius):
    combinations = []
    for i in range(1, outerRadius+1):
        combinations.extend("".join(p) for p in itertools.product(monomers, repeat=i))
    return combinations

def formatFile(path):
    with open(path, "r") as f:
        lines = f.readlines()
    lines = [l for l in lines if l.strip()]
    species, food, reactions = [], [], []
    for line in lines:
        if "SEED" in line:
            continue
        if "+" not in line and not NUMERIC_START.match(line):
            line = line.split()
            line = str(f"{line[0]:<15}{line[1]:<30}{line[2]:<10}\n")
            species.append(line)
        elif "+" in line and not NUMERIC_START.match(line):
            if "FC" in line:
                line = line.split()
                line = str(f"{line[0]} {line[1]} {line[2]} {line[3]} {line[4]} {line[5]} {line[6]} {line[7]} {line[8]}\n")
            else:
                line = line.split()
                line = str(f"{line[0]} {line[1]} {line[2]} {line[3]} {line[4]} {line[5]} {line[6]} {line[7]} {line[8]} {line[9]} {line[10]}\n")
            reactions.append(line)
        elif NUMERIC_START.match(line):
            line = line.split()
            line = str(f"{line[0]} {line[1]} {line[2]} {line[3]} {line[4]}\n")
            food.append(line)
    with open(path, "w") as f:
        species.append("\n")
        f.writelines(species + food + reactions)

def getSpeciesNamesFromFile(path):
    species = []
    with open(path, "r") as f:
        lines = f.readlines()
    for line in lines:
        if "+" not in line and not NUMERIC_START.match(line) and line.strip():
            line = line.split()
            species.append(line[0])
    return species

def getSpeciesCountFromFile(path):
    count = 0
    with open(path, "r") as f:
        lines = f.readlines()
    for line in lines:
        if "+" not in line and not NUMERIC_START.match(line) and line.strip():
            count += 1
    return count

def getNotNullSpeciesCountFromFile(path):
    count = 0
    with open(path, "r") as f:
        lines = f.readlines()
    for line in lines:
        if "+" not in line and not NUMERIC_START.match(line) and line.strip():
            line = line.split()
            if float(line[1]) > 0:
                count += 1
    return count

def getFoodsNamesFromFile(path):
    foods = []
    with open(path, "r") as f:
        lines = f.readlines()
    for line in lines:
        if NUMERIC_START.match(line):
            line = line.split()
            foods.append(line[2])
    return foods

def runPitzalisContainers(PITZALIS):
    os.system(f'cd "{PITZALIS}/src/Utils" && python ./t_Carpi_Pitzalis.py > nul 2>&1')

def runPitzalisSimulator(PITZALIS):
    os.system(f'cd "{PITZALIS}/src" && python ./main.py -v')

def getSpeciesConcentrations(path):
    concentrationsDictionary = {}
    data = []
    with open(path, "r") as f:
        lines = f.readlines()
    lines = [l for l in lines if l.strip()]
    for line in lines:
        line = line.split()
        data.append(line)
    species = data[0]
    quantity = data[1]
    for i in range(len(species)):
        concentration = float(quantity[i])
        if "Cont" in species[i]:
            concentration *= 0.5
        else:
            concentration *= 0.353553
        concentrationsDictionary[species[i]] = concentration
    return concentrationsDictionary

def runMutator():
    os.system(f'python ./main.py -d mutate')

# Random functions
def generateSeed(seed):
    if seed:
        random.seed(seed)
    else:
        seed=getRandomValue()
        random.seed(seed)
    return seed

def calculateProbability(probability):
    return (random.random() < probability)

def calculateRandomValue(minValue, maxValue):
    return random.randint(minValue, maxValue)

def calculateActiveSiteLength(catalyst, minActiveSiteLength, maxActiveSiteLength):
    length=calculateRandomValue(minActiveSiteLength, maxActiveSiteLength)
    while len(catalyst.getName()) < length:
        length=calculateRandomValue(minActiveSiteLength, maxActiveSiteLength)
    return length

# Aesthetic functions
def underlinedHeader(headers):
    headers = [underlinedPrint(header) for header in headers]
    return headers

def underlinedPrint(text):
    return (colored(text,attrs=['underline']))

def boldTitle(text):
    return (colored(text,'blue',attrs=['bold','underline']))

def greyText(text):
    return (colored(text,'grey'))

def boldGreenTitle(text):
    return (colored(text,'green',attrs=['bold','underline']))

def error(text):
    return (colored(text,'red',attrs=['bold']))

# Debug print functions
def printParameters(parameters,kauffman=False):
    print(boldTitle("\nPARAMETRI"))
    if not parameters:
        print(error("NESSUN PARAMETRO PRESENTE"))
        return
    if kauffman:
        generatorParam=["lowerLimitForCatalyst","maxGenerationTime","maxMembraneLength","maxProductLength","maxReactionProduced","monomers","outputFile","probabilityOfCleavage","seed"]
        printSelectedParameters(parameters, generatorParam, kauffman)
    else:
        generatorParam=["initialCleavageCatalysts","initialCondensationCatalysts","lowerLimitForCatalyst","maxActiveSiteLength","maxCatalystLength","maxCleavageLength","maxCondensationLength","maxGenerationTime","minActiveSiteLength","monomers","outputFile","outputRulesFile","probabilityOfCatalyst","probabilityOfCleavage","seed"]
        printSelectedParameters(parameters, generatorParam)

def printSelectedParameters(parameters, selectedParameters,kauffman=False):
    headers = ["PARAMETRO", "VALORE"]
    table=[]
    for p in selectedParameters:
        if p == "seed" and parameters[p] == None:
            table.append([p,"None"])
        elif p == "monomers" and not kauffman:
            str=""
            for m in parameters[p]:
                str += "Nome: "+m["name"]+" "+"Complementare: "+m["complementary"]+"\n"
            table.append([p,str])
        elif p == "monomers" and kauffman:
            str=""
            for m in parameters[p]:
                str += "Nome: "+m+"\n"
            table.append([p,str])
        else:
            table.append([p,parameters[p]])
    print(tabulate(table, underlinedHeader(headers), tablefmt="simple_grid"))

def printSpecies(species,ended=False,kauffman=False):
    if not species:
        print(error("NESSUNA SPECIE PRESENTE"))
        return
    if not ended:
        print(boldTitle("SPECIE"))
    else:
        print(boldGreenTitle("SPECIE GENERATE"))
    if kauffman:
        for s in species:
            print("• "+ s.getName())
    else:
        headers = ["SPECIE", "INIZIALI"]
        table = [[s.name, "Sì" if s.isInitial else "No"] for s in species]
        print(tabulate(table, underlinedHeader(headers), tablefmt="simple_grid"))

def printReactionClasses(reactionClasses,new=False):
    if not reactionClasses:
        print(error("NESSUNA CLASSE DI REAZIONE PRESENTE"))
        return
    headers = ["SPECIE", "TIPO", "SITO ATTIVO", "POSIZIONE", "CLASSE"]
    table = []
    for rc in reactionClasses:
        if rc.getCatalyst().getIsCondensation():
            table.append([rc.getCatalyst().name, 'Condensazione', rc.getCatalyst().getName()[rc.getStart():rc.getEnd()],rc.getSplit(),'R-'+rc.getReagents()[0]+'  '+rc.getReagents()[1]+'-R'])
        else:
            table.append([rc.getCatalyst().name, 'Cleavage', rc.getCatalyst().getName()[rc.getStart():rc.getEnd()],rc.getSplit(), 'R-'+rc.getReagents()[0]+'-R'])
    if not new: 
        print(boldTitle("CLASSI DI REAZIONI"))
    else: 
        print(boldTitle("NUOVE CLASSI GENERATE"))
    print(tabulate(table, underlinedHeader(headers), tablefmt="simple_grid"))

def printReactions(reactions, kauffman=False):
    if not reactions:
        print(error("NESSUNA REAZIONE PRESENTE"))
        print(error("PER GENERARE LE REAZIONI MODIFICARE I PARAMETRI"))
        return
    print(boldGreenTitle("REAZIONI GENERATE"))
    if kauffman:
        for r in reactions:
            print(r.printReaction(kauffman=True))
    else:
        for r in reactions:
            print(r.printReaction())
    

def orderSpecies(species):
        species = sorted(species, key=lambda x: x.getLength())
        return species

def printTabulatedCatalysts(catalysts):
    for c in catalysts:
        print(c.getName(), c.getLength(), c.getTotalRules(), c.getCondensationRules(), c.getCleavageRules(), c.getTotalReactions(), c.getCondensationReactions(), c.getCleavageReactions(), c.getCatalyzerAsReagent(), c.getNumberOfCatalyzedSpecies(), c.getCatalyzedSpecies())

def printTabulatedSpecies(species):
        for s in species:
            print(s.getName(), s.getLength(), s.getTotalProducts(), s.getCondensationProducts(), s.getCleavageProducts(), s.getTotalCatalyzers(), s.getCondensationCatalyzers(), s.getCleavageCatalyzers(), s.getCatalyzers(), s.getSpeciesAsReactar())

# Output print functions
def deleteReportFile(kauffman=False,mutator=False):
    if mutator:
        function="mutator"
    else:
        if not kauffman:
            function="generator"
        else:
            function="kauffmanGenerator"
    file=os.path.join(BASE_DIR, "io/"+function+"/output/report.txt")
    if os.path.exists(file):
        os.remove(file)

def writeReportFile(parameters,data,kauffman=False,mutator=False):
    if mutator:
        function="mutator"
    else:
        if not kauffman:
            function="generator"
        else:
            function="kauffmanGenerator"
    file=os.path.join(BASE_DIR, "io/"+function+"/output/report.txt")
    with open(file, 'w') as f:
        f.write("PARAMETRI\n")
        for key in parameters:
            f.write(f"{key}: {parameters[key]}\n")
        f.write("\n")
        type=data["error"]
        f.write("DATI\n")
        if kauffman:
            f.write("Numero di reazioni generate correttamente: "+str(data["lap"])+"\n")
            f.write("Numero di reazioni mancanti: "+str(parameters["maxReactionProduced"]-data["lap"])+"\n")
        else:
            if type=="START":
                f.write("Numero di giri svolti completamente: "+str(data["lap"])+"\n")
                f.write("Non sono stati interrotti i processi di sviluppo di nuove reazioni o specie\n")
            if type=="REACTION":
                f.write("Numero di giri svolti completamente: "+str(data["lap"])+"\n")
                f.write("Generazione interrotta al giro "+str(data["lap"]+1)+" durante lo sviluppo di nuove reazioni\n")
                remainingReactions = len(data["reactionClasses"])-len(data["processedReactionClasses"])
                f.write("Ci sono ancora "+str(remainingReactions)+" classi di reazioni da analizzare: \n")
                for rc in data["reactionClasses"]:
                    if rc in data["processedReactionClasses"]:
                        f.write(rc.printReactionClass()+"\n")
                    else:
                        f.write(rc.printReactionClass()+" <--\n")
            if type=="SPECIES":
                f.write("Numero di giri svolti completamente: "+str(data["lap"])+"\n")
                f.write("Generazione interrotta al giro "+str(data["lap"]+1)+" durante la ricerca di nuove specie\n")
                remainingReactions = len(data["reactions"])-len(data["processedReactions"])
                f.write("Ci sono ancora "+str(remainingReactions)+" reazioni da analizzare: \n")
                for r in data["reactions"]:
                    if r in data["processedReactions"]:
                        f.write(r.printReaction()+"\n")
                    else:
                        f.write(r.printReaction()+" <--\n")
                

def writeOutputFile(seed,parameters,species,allReactions,uniqueReactions,kauffman=False,mutator=False):
    if mutator:
        function="mutator"
        name = ""
        uniqueReactionsFile = os.path.join(BASE_DIR, "io/"+function+"/output/"+name+"uniqueReactions.txt")
        allReactionsFile = None
        multiplicyReactionsFile = None
        files = [uniqueReactionsFile]
    else:
        name = parameters['outputFile']+"-"
        if not kauffman:
            function="generator"
        else:
            function="kauffmanGenerator"
        uniqueReactionsFile = os.path.join(BASE_DIR, "io/"+function+"/output/"+name+"uniqueReactions.txt")
        allReactionsFile = os.path.join(BASE_DIR, "io/"+function+"/output/"+name+"allReactions.txt")
        multiplicyReactionsFile = os.path.join(BASE_DIR, "io/"+function+"/output/"+name+"multiplicyReactions.txt")
        files = [allReactionsFile, uniqueReactionsFile, multiplicyReactionsFile]
    for file in files:
        with open(file, 'w') as f:
            f.write(f"SEED UTILIZZATO: {seed}\n\n")
            f.write((f"{"Cont":<15}{'1.35e-16':<10}{"0.0":<10}\n"))
            sortedSpecies=sorted(species,key=len)
            crossMembraneSpecies=[]
            for s in sortedSpecies:
                if not s.getIsFood():
                    val="0.01"
                else:
                    val="0.0"
                    crossMembraneSpecies.append(s)
                f.write(f"{s.getName():<15}{parameters['internalConcentration']:<10}{val:<10}\n")
            f.write("\n")
            for s in crossMembraneSpecies:
                f.write(f"{10.0} > {s.getName()} ; {parameters['internalConcentration']}\n")
            if file == allReactionsFile and allReactionsFile is not None:   
                if not allReactions:
                    f.write("NESSUNA REAZIONE GENERATA, SI CONSIGLIA DI MODIFICARE I PARAMETRI\n")
                else:
                    if kauffman:
                        for r in allReactions:
                            f.write(f"{r.printReaction(kauffman=True)} ; {parameters['reactionCoefficient']}"+ "\n") 
                    else:
                        for r in allReactions:
                            f.write(f"{r.printReaction()} ; {parameters['reactionCoefficient']}"+ "\n") 
            if file == uniqueReactionsFile:
                if not uniqueReactions:
                    f.write("NESSUNA REAZIONE GENERATA, SI CONSIGLIA DI MODIFICARE I PARAMETRI\n")
                else:
                    if kauffman:
                        for r in uniqueReactions:
                            f.write(f"{r.printReaction(kauffman=True)} ; {parameters['reactionCoefficient']}"+ "\n") 
                    else:
                        for r in uniqueReactions:
                            f.write(f"{r.printReaction()} ; {parameters['reactionCoefficient']}"+ "\n") 
            if file == multiplicyReactionsFile and multiplicyReactionsFile is not None:
                if not uniqueReactions:
                    f.write("NESSUNA REAZIONE GENERATA, SI CONSIGLIA DI MODIFICARE I PARAMETRI\n")
                else:
                    if kauffman:
                        for r in uniqueReactions:
                            multiplicity = r.getMultiplicity()
                            f.write(f"{str(multiplicity)+"x":<2} {r.printReaction(kauffman=True)} ; {parameters['reactionCoefficient']}"+ "\n") 
                    else:
                        for r in uniqueReactions:
                            multiplicity = r.getMultiplicity()
                            f.write(f"{str(multiplicity)+"x":<2} {r.printReaction()} ; {parameters['reactionCoefficient']}"+ "\n") 

def writeRulesFile(parameters,reactionClasses,mutator=False):
    if mutator:
        path = os.path.join(BASE_DIR, "io/mutator/output/rules.txt")
    else:
        path = os.path.join(BASE_DIR, "io/generator/output/"+parameters['outputRulesFile'])
    with open(path, 'w') as f:
        f.write(f"{"CATALIZZATORE":<15} {"TIPO":<17} {"SITO ATTIVO":<13} {"POSIZIONE":<11} {"CLASSE":<10}\n")
        for rc in reactionClasses:
            if rc.getCatalyst().getIsCondensation():
                f.write(f"{rc.getCatalyst().getName():<15} {'Condensazione':<17} {rc.getCatalyst().getName()[rc.getStart():rc.getEnd()]:<13} {rc.getSplit():<11} {'R-'+rc.getReagents()[0]+'+'+rc.getReagents()[1]+'-R':<10}\n")
            else:
                f.write(f"{rc.getCatalyst().getName():<15} {'Cleavage':<17} {rc.getCatalyst().getName()[rc.getStart():rc.getEnd()]:<13} {rc.getSplit():<11} {'R-'+rc.getReagents()[0][:rc.getSplit()]+' '+rc.getReagents()[0][rc.getSplit():]+'-R':<10}\n")

def duplicateFilesForTabulator(parameters,kauffman=False,mutator=False):
    if mutator:
        shutil.copy(os.path.join(BASE_DIR,"io/mutator/output/uniqueReactions.txt"),os.path.join(BASE_DIR,"io/tabulator/input/chemistry.txt"))
        shutil.copy(os.path.join(BASE_DIR,"io/mutator/output/rules.txt"),os.path.join(BASE_DIR,"io/tabulator/input/chemistryRules.txt"))
    else:
        if not kauffman:
            shutil.copy(os.path.join(BASE_DIR,"io/generator/output/"+parameters['outputFile']+"-uniqueReactions.txt"),os.path.join(BASE_DIR,"io/tabulator/input/chemistry.txt"))
            shutil.copy(os.path.join(BASE_DIR,"io/generator/output/"+parameters['outputRulesFile']),os.path.join(BASE_DIR,"io/tabulator/input/chemistryRules.txt"))
        else:
            if os.path.exists(os.path.join(BASE_DIR,"io/tabulator/input/chemistryRules.txt")):
                os.remove(os.path.join(BASE_DIR,"io/tabulator/input/chemistryRules.txt"))
            shutil.copy(os.path.join(BASE_DIR,"io/kauffmanGenerator/output/"+parameters['outputFile']+"-uniqueReactions.txt"),os.path.join(BASE_DIR,"io/tabulator/input/chemistry.txt"))

def cleanReaction(reaction):
    formatted = re.sub(r'[^A-Za-z+]','',reaction)
    return formatted.split('+')

def setTable(pd, rows, columns):
    df = pd.DataFrame(rows, columns=columns)
    return df

def setTitle(ws, title, startRow, startCol, endCol):
    ws[f"{get_column_letter(startCol)}{startRow}"] = title
    ws[f"{get_column_letter(startCol)}{startRow}"].font = Font(size=14, bold=True)
    ws[f"{get_column_letter(startCol)}{startRow}"].alignment = Alignment(horizontal="center")
    ws[f"{get_column_letter(startCol)}{startRow}"].fill = PatternFill(start_color="d3d3d3", end_color="d3d3d3", fill_type="solid")
    ws[f"{get_column_letter(startCol)}{startRow}"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells(start_row=startRow, start_column=startCol, end_row=startRow, end_column=endCol)

def setTableStyle(ws, start_row, columns, df, border, color1, color2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=start_row+len(df), min_col=2, max_col=2+len(columns)-1), start=1):
        if i % 2 != 0:
            fill = color1 
        else:
            fill = color2
        for cell in row:
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.alignment = Alignment(vertical='center')

def resizeCells(ws, titles, catalystsColumns, speciesColumns):
    for col in ws.iter_cols():
        maxLength = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value and isinstance(cell.value, str):
                if cell.value in titles:
                    continue
                if cell.value in catalystsColumns or cell.value in speciesColumns:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                maxLength = max(maxLength, len(cell.value))
        width = min(maxLength+2, 30)
        ws.column_dimensions[column].width = width

def writeOnExcelFile(catalysts,species):
    catalystColumns = ['Name', 'Length', 'Reaction class', 'Condensation class', 'Cleavage class', 'Total generated reactions', 'Generated condensation reactions', 'Generated cleavage reactions', 'Catalyzers as reagent', 'Species catalized', 'Nr. species catalized']
    catalystRows = []
    for c in catalysts:
        catalystRows.append([c.getName(), c.getLength(), c.getTotalRules(), c.getCondensationRules(),c.getCleavageRules(), c.getTotalReactions(), c.getCondensationReactions(),c.getCleavageReactions(), c.getCatalyzerAsReagent(), c.getCatalyzedSpecies(), c.getNumberOfCatalyzedSpecies()])
    catalystDf = setTable(pd, catalystRows, catalystColumns)

    speciesColumns = ['Name', 'Length', 'Reaction as product', 'Condensation as products', 'Cleavage as products', 'Total catalyzers', 'Condensation catalyzers', 'Cleavage catalyzers', 'Catalyzers', 'Reactions as reactants']
    speciesRows = []
    for s in species:
        speciesRows.append([s.getName(), s.getLength(), s.getTotalProducts(), s.getCondensationProducts(),s.getCleavageProducts(), s.getTotalCatalyzers(), s.getCondensationCatalyzers(),s.getCleavageCatalyzers(), s.getCatalyzers(), s.getSpeciesAsReactar()])
    speciesDf = setTable(pd, speciesRows, speciesColumns)

    catalyzerAsSpeciesRows=[]
    for c in catalysts:
        for s in species:
            if c.getName() == s.getName():
                catalyzerAsSpeciesRows.append([s.getName(), s.getLength(), s.getTotalProducts(), s.getCondensationProducts(),s.getCleavageProducts(), s.getTotalCatalyzers(), s.getCondensationCatalyzers(),s.getCleavageCatalyzers(), s.getCatalyzers(), s.getSpeciesAsReactar()])
    catalyzersAsSpeciesDf = setTable(pd, catalyzerAsSpeciesRows, speciesColumns)

    path = os.path.join(BASE_DIR, "io/Tabulator/output/output.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        catalystDf.to_excel(writer, index=False, startrow=2, startcol=1, sheet_name="Sheet1")
        speciesDf.to_excel(writer, index=False, startrow=len(catalystDf)+5, startcol=1, sheet_name="Sheet1")
        catalyzersAsSpeciesDf.to_excel(writer, index=False, startrow=len(catalystDf)+len(speciesDf)+8, startcol=1, sheet_name="Sheet1")

    wb = load_workbook(path)
    ws = wb.active
    titles = ["Catalysts Information", "Species Information", "Catalyzers as Species Information"]
    setTitle(ws, titles[0], startRow=2, startCol=2, endCol=len(catalystColumns)+1)
    speciesSR = len(catalystDf)+5
    setTitle(ws, titles[1], startRow=speciesSR, startCol=2, endCol=len(speciesColumns)+1)
    catalyzersAsSpeciesSR = speciesSR+len(speciesDf)+3
    setTitle(ws, titles[2], startRow=catalyzersAsSpeciesSR, startCol=2, endCol=len(speciesColumns)+1)

    white = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    lightGray = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    setTableStyle(ws, 3, catalystColumns, catalystDf, border, white, lightGray)
    setTableStyle(ws, len(catalystDf)+6, speciesColumns, speciesDf, border, white, lightGray)
    setTableStyle(ws, len(catalystDf)+len(speciesDf)+9, speciesColumns, catalyzersAsSpeciesDf, border, white, lightGray)
    resizeCells(ws,titles,catalystColumns,speciesColumns)

    wb.save(path)

def writeAnalysOnExcel(data, serieName):
    columns = ['Lap', 'Interrupted', 'All species', 'All reactions', 'Unique species', 'Unique reactions','Membrane reactions', 'Catalysts', 'RAF', 'RAF species', 'RAF reactions', 'RAF membrane reactions', 'RAF catalysts']
    rows = []
    for r in data:
        if not r[1] :
            r[1] = "No"
        else:
            r[1] = "Yes"
        if not r[8]:
            r[8] = "No"
        else:
            r[8] = "Yes"
        rows.append(r)

    analystDf = setTable(pd, rows, columns)

    path = os.path.join(BASE_DIR, "io", "launcher", "output", "series", serieName, "analysis.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        analystDf.to_excel(writer, index=False, startrow=2, startcol=1, sheet_name="Sheet1")

    wb = load_workbook(path)
    ws = wb.active
    setTitle(ws, "Analysis", startRow=2, startCol=2, endCol=len(columns)+1)

    white = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    lightGray = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    setTableStyle(ws, 3, columns, analystDf, border, white, lightGray)
    resizeCells(ws, ["Analysis"], columns, columns)
    wb.save(path)

def writeEvolverAnalysis(generations, countSpecies, countNotNullSpecies, timeRecords, acceptedStatus, directory):
    columns = ['Generation', 'Number of species', 'Number of non-zero species', 'Time', 'Accepted']
    rows = []
    for i in range(generations):
        if acceptedStatus[i]:
            status = "Yes"
        else:
            status = "No"
        rows.append([i, countSpecies[i], countNotNullSpecies[i], timeRecords[i], status])

    analystDf = setTable(pd, rows, columns)

    path = os.path.join(BASE_DIR, "io", "evolver", "output", directory, "analysis.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        analystDf.to_excel(writer, index=False, startrow=2, startcol=1, sheet_name="Sheet1")

    wb = load_workbook(path)
    ws = wb.active
    setTitle(ws, "Evolver Analysis", startRow=2, startCol=2, endCol=len(columns)+1)

    white = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    lightGray = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    setTableStyle(ws, 3, columns, analystDf, border, white, lightGray)
    resizeCells(ws, ["Evolver Analysis"], columns, columns)
    wb.save(path)